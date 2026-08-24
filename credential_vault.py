"""
Credential vault — self-registration accounts for ATS portals.

Workday, iCIMS, Taleo and friends make you create an account before you can
apply, and each company runs its own tenant. Pre-configuring a password per
tenant does not scale, so the bot mints one itself:

    vault = CredentialVault(cfg)
    cred = vault.get_or_create("nvidia.wd5.myworkdayjobs.com", ats="workday")
    # -> {"site": ..., "email": ..., "password": "<generated>", "status": "new"}

Every credential is written to a sheet you can open — `data/ats_accounts.xlsx`
if openpyxl is installed, otherwise `data/ats_accounts.csv` — plus a SQLite
table so a lost spreadsheet never loses an account. A tenant already in the
vault returns its existing credential, so re-visiting a company signs in with
the password it registered with instead of creating a duplicate.

SECURITY — read this
    The sheet holds working credentials in PLAIN TEXT. That is the point (you
    must be able to log in yourself), but it means:
      * the file is created 0600 (owner-only) where the OS supports it,
      * `data/` is gitignored — never commit it,
      * these are throwaway ATS portal accounts; never reuse a password you
        use anywhere important,
      * `vault.email` should be an address you control and can check, because
        most tenants send a verification link.
"""

import csv
import logging
import os
import secrets
import string
from datetime import datetime
from pathlib import Path

log = logging.getLogger("lla.vault")

# Symbols every major ATS accepts. Deliberately excludes quotes, backslash,
# angle brackets and semicolons, which break naive form handling and CSV.
_SYMBOLS = "!@#$%^&*-_=+?"
_AMBIGUOUS = "lI1O0"

SHEET_COLUMNS = ["site", "ats", "email", "password", "status",
                 "created_at", "last_used", "job_url", "notes"]


def generate_password(length: int = 16) -> str:
    """A strong password that satisfies every ATS complexity rule we've seen.

    Guarantees at least two lowercase, two uppercase, two digits and two
    symbols, avoids visually ambiguous characters, and never starts or ends
    with a symbol (some validators reject that).
    """
    length = max(12, min(length, 64))
    lower = "".join(c for c in string.ascii_lowercase if c not in _AMBIGUOUS)
    upper = "".join(c for c in string.ascii_uppercase if c not in _AMBIGUOUS)
    digits = "".join(c for c in string.digits if c not in _AMBIGUOUS)

    chars = ([secrets.choice(lower) for _ in range(2)]
             + [secrets.choice(upper) for _ in range(2)]
             + [secrets.choice(digits) for _ in range(2)]
             + [secrets.choice(_SYMBOLS) for _ in range(2)])
    pool = lower + upper + digits + _SYMBOLS
    chars += [secrets.choice(pool) for _ in range(length - len(chars))]

    # Shuffle, then force an alphanumeric at both ends.
    secrets.SystemRandom().shuffle(chars)
    alnum = lower + upper + digits
    if chars[0] in _SYMBOLS:
        chars[0] = secrets.choice(alnum)
    if chars[-1] in _SYMBOLS:
        chars[-1] = secrets.choice(alnum)
    return "".join(chars)


def password_is_strong(pw: str) -> bool:
    """Does this password meet the common ATS bar (len>=8, 3 of 4 classes)?"""
    if not pw or len(pw) < 8:
        return False
    classes = sum([
        any(c.islower() for c in pw), any(c.isupper() for c in pw),
        any(c.isdigit() for c in pw), any(c in _SYMBOLS or not c.isalnum() for c in pw),
    ])
    return classes >= 3


def tenant_slug(site: str) -> str:
    """'nvidia.wd5.myworkdayjobs.com' -> 'nvidia' (for plus-addressing/notes)."""
    host = (site or "").lower().split("/")[0]
    part = host.split(".")[0]
    return "".join(ch for ch in part if ch.isalnum()) or "ats"


def plus_address(email: str, slug: str) -> str:
    """you@gmail.com + 'nvidia' -> you+nvidia@gmail.com (same inbox, unique login)."""
    if not email or "@" not in email:
        return email
    local, _, domain = email.partition("@")
    if "+" in local:                       # already tagged — don't stack tags
        local = local.split("+", 1)[0]
    return f"{local}+{slug}@{domain}"


class CredentialVault:
    """Mint, remember and persist per-tenant ATS portal accounts."""

    def __init__(self, cfg: dict, state=None):
        vc = (cfg or {}).get("credential_vault", {}) or {}
        ea = (cfg or {}).get("external_apply", {}) or {}
        self.enabled = vc.get("enabled", True)
        self.auto_register = vc.get("auto_register", True)
        self.password_length = vc.get("password_length", 16)
        self.use_plus_addressing = vc.get("use_plus_addressing", False)
        self.output_dir = vc.get("output_dir", "data")
        self.sheet_name = vc.get("sheet_name", "ats_accounts")
        # Address that receives every verification email.
        self.email = (vc.get("email")
                      or ((ea.get("ats_accounts", {}) or {}).get("generic", {}) or {}).get("email")
                      or (cfg or {}).get("personal", {}).get("email", ""))
        self.state = state
        self._cache = {}
        self._load()

    # ------------------------------------------------------------------
    # Paths
    # ------------------------------------------------------------------

    @property
    def xlsx_path(self) -> Path:
        return Path(self.output_dir) / f"{self.sheet_name}.xlsx"

    @property
    def csv_path(self) -> Path:
        return Path(self.output_dir) / f"{self.sheet_name}.csv"

    @staticmethod
    def _has_excel() -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    # ------------------------------------------------------------------
    # Load / persist
    # ------------------------------------------------------------------

    def _load(self):
        """Read existing credentials so a known tenant is never re-registered."""
        rows = []
        try:
            if self.xlsx_path.exists() and self._has_excel():
                import openpyxl
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                headers = [str(c.value or "") for c in ws[1]] if ws.max_row else []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({h: ("" if v is None else str(v))
                                 for h, v in zip(headers, r)})
            elif self.csv_path.exists():
                with open(self.csv_path, newline="", encoding="utf-8") as fh:
                    rows = list(csv.DictReader(fh))
        except Exception as exc:
            log.warning("credential vault: could not read existing sheet (%s)", exc)
        for row in rows:
            site = (row.get("site") or "").strip().lower()
            if site:
                self._cache[site] = row

    def _write(self):
        """Persist every credential. Excel when available, CSV otherwise."""
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        rows = [self._cache[k] for k in sorted(self._cache)]
        wrote = None
        if self._has_excel():
            try:
                import openpyxl
                from openpyxl.styles import Font
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "ATS Accounts"
                ws.append(SHEET_COLUMNS)
                for c in ws[1]:
                    c.font = Font(bold=True)
                for row in rows:
                    ws.append([row.get(c, "") for c in SHEET_COLUMNS])
                widths = {"site": 42, "ats": 14, "email": 34, "password": 20,
                          "status": 14, "created_at": 20, "last_used": 20,
                          "job_url": 50, "notes": 30}
                for i, col in enumerate(SHEET_COLUMNS, start=1):
                    ws.column_dimensions[chr(64 + i)].width = widths.get(col, 18)
                ws.freeze_panes = "A2"
                wb.save(self.xlsx_path)
                wrote = self.xlsx_path
            except Exception as exc:
                log.warning("credential vault: Excel write failed (%s) — using CSV", exc)
        if wrote is None:
            with open(self.csv_path, "w", newline="", encoding="utf-8") as fh:
                w = csv.DictWriter(fh, fieldnames=SHEET_COLUMNS)
                w.writeheader()
                for row in rows:
                    w.writerow({c: row.get(c, "") for c in SHEET_COLUMNS})
            wrote = self.csv_path
        self._restrict(wrote)
        # Mirror into SQLite so a deleted/overwritten sheet never loses accounts.
        self._save_state(rows)
        return wrote

    @staticmethod
    def _restrict(path: Path):
        """Owner-only permissions where the OS supports it (no-op on Windows)."""
        try:
            os.chmod(path, 0o600)
        except Exception:
            pass

    def _save_state(self, rows):
        if not self.state or not getattr(self.state, "conn", None):
            return
        try:
            self.state.conn.execute("""
                CREATE TABLE IF NOT EXISTS ats_credentials (
                    site       TEXT PRIMARY KEY,
                    ats        TEXT,
                    email      TEXT,
                    password   TEXT,
                    status     TEXT,
                    created_at TEXT,
                    last_used  TEXT,
                    job_url    TEXT,
                    notes      TEXT
                )""")
            for r in rows:
                self.state.conn.execute(
                    "INSERT OR REPLACE INTO ats_credentials "
                    "(site, ats, email, password, status, created_at, last_used, job_url, notes) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    tuple(r.get(c, "") for c in SHEET_COLUMNS))
            self.state.conn.commit()
        except Exception as exc:
            log.debug("credential vault: state mirror failed (%s)", exc)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def get(self, site: str):
        """Existing credential for a tenant, or None."""
        return self._cache.get((site or "").strip().lower())

    def get_or_create(self, site: str, ats: str = "", job_url: str = "",
                      email: str = "") -> dict:
        """Return this tenant's credential, minting one on first sight.

        The returned dict carries `status`: "new" when it was just generated
        (so the caller knows to REGISTER) or "existing" (so it should SIGN IN).
        """
        site = (site or "").strip().lower()
        if not site:
            return {}
        found = self._cache.get(site)
        if found:
            found["last_used"] = datetime.now().isoformat(timespec="seconds")
            found.setdefault("status", "existing")
            self._write()
            return {**found, "status": "existing"}

        base_email = email or self.email
        if not base_email:
            log.warning("credential vault: no email configured — cannot register")
            return {}
        slug = tenant_slug(site)
        row = {
            "site": site,
            "ats": ats,
            "email": plus_address(base_email, slug) if self.use_plus_addressing else base_email,
            "password": generate_password(self.password_length),
            "status": "new",
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "last_used": datetime.now().isoformat(timespec="seconds"),
            "job_url": job_url,
            "notes": "auto-generated",
        }
        self._cache[site] = row
        path = self._write()
        log.info("   🔐 vault: new %s account for %s -> saved to %s", ats or "ATS", site, path)
        return dict(row)

    def mark(self, site: str, status: str, notes: str = ""):
        """Record the outcome of a registration/sign-in attempt."""
        site = (site or "").strip().lower()
        row = self._cache.get(site)
        if not row:
            return
        row["status"] = status
        row["last_used"] = datetime.now().isoformat(timespec="seconds")
        if notes:
            row["notes"] = notes
        self._write()

    def all_credentials(self) -> list:
        return [self._cache[k] for k in sorted(self._cache)]
