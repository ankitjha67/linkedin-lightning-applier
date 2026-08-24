"""
Workday self-registration E2E — run against the bundled replica page.

    cd tests/e2e
    python3 -m http.server ... (or use serve_fixtures.py)
    python3 workday_registration_e2e.py

Serves fixtures/workday_account.html over local HTTPS as a Workday tenant, then
drives the REAL WorkdayHandler + CredentialVault: registers with a generated
password, writes the accounts sheet, and re-visits to prove it signs in instead
of registering twice.

End-to-end: the REAL WorkdayHandler registering on a Workday-replica account
page in a REAL browser, with a REAL credential vault writing a real xlsx.

Nothing is mocked except the tenant hostname (mapped to a local HTTPS server).
"""
import os, shutil, sys, tempfile

os.environ["PATH"] = ":".join(p for p in os.environ["PATH"].split(":") if "node22" not in p)
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from ats_handlers import get_handler
from credential_vault import CredentialVault

TENANT = "nvidia.wd5.myworkdayjobs.com"
URL = f"https://{TENANT}/wd/account"
OUT = tempfile.mkdtemp()

def browser():
    o = Options()
    if os.environ.get("LLA_E2E_CHROME"):
        o.binary_location = os.environ["LLA_E2E_CHROME"]
    for a in ("--headless=new", "--no-sandbox", "--disable-dev-shm-usage",
              "--ignore-certificate-errors", "--no-proxy-server",
              f"--host-resolver-rules=MAP {TENANT} 127.0.0.1:8443"):
        o.add_argument(a)
    return webdriver.Chrome(options=o)

CFG = {
    "personal": {"email": "candidate@example.com", "first_name": "Ankit",
                 "last_name": "Kumar", "city": "Gurugram", "country": "India"},
    "credential_vault": {"output_dir": OUT},
    "external_apply": {"ats_accounts": {}},   # nothing pre-configured on purpose
}

fails = []
def check(cond, msg):
    print(f"  {'✅' if cond else '🐞'} {msg}")
    if not cond:
        fails.append(msg)

print("\n━━ 1. FIRST VISIT — autonomous registration (nothing pre-configured)")
d = browser()
try:
    d.get(URL)
    h = get_handler("workday", None, CFG)
    check(h._needs_auth(d), "auth wall detected on the create-account page")

    # The live password-policy hint must NOT read as an error (bug 5).
    check(not h._auth_error(d), "password-policy hint is not treated as an error")

    ok = h._authenticate(d, job_url="https://x/job/1")
    check(ok, "_authenticate() returned True (registered)")

    reg = d.execute_script("return window.__registered || null;")
    check(bool(reg), f"page confirms registration: {reg}")
    if reg:
        check(reg["email"] == "candidate@example.com", "registered with the configured email")
        pw = reg["password"]
        strong = (len(pw) >= 12 and any(c.islower() for c in pw)
                  and any(c.isupper() for c in pw) and any(c.isdigit() for c in pw)
                  and any(not c.isalnum() for c in pw))
        check(strong, f"generated password is strong (len={len(pw)})")

    # Bug 4: page still says "Sign In" in the header, but auth is done.
    check("Sign In" in d.page_source, "page still contains 'Sign In' (the bug-4 trap)")
    check(h._auth_complete(d), "_auth_complete() correctly True after success")
finally:
    d.quit()

print("\n━━ 2. THE SHEET — credentials saved where the user can read them")
v = CredentialVault(CFG)
cred = v.get(TENANT)
check(bool(cred), "credential row exists for the tenant")
if cred:
    check(cred["status"] == "registered", f"status recorded = {cred['status']!r}")
    check(cred["password"] == (reg or {}).get("password"),
          "sheet password matches what was actually submitted to the site")
    check(cred["job_url"] == "https://x/job/1", "job url recorded")
sheet = v.xlsx_path if v.xlsx_path.exists() else v.csv_path
check(sheet.exists(), f"sheet written: {sheet.name}")
if v.xlsx_path.exists():
    import openpyxl
    ws = openpyxl.load_workbook(v.xlsx_path).active
    hdr = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    print(f"     columns: {hdr}")
    print(f"     row    : {row[0]} | {row[2]} | {row[3][:6]}…({len(row[3])} chars) | {row[4]}")
    check("password" in hdr and row[3], "xlsx contains a readable password column")

print("\n━━ 3. SECOND VISIT — must SIGN IN, not register a duplicate")
d2 = browser()
try:
    d2.get(URL)
    h2 = get_handler("workday", None, CFG)
    acct = h2._account(d2)
    check(acct["status"] == "existing", "vault reports the tenant as already registered")
    check(acct["password"] == cred["password"], "reuses the SAME stored password")
finally:
    d2.quit()

print("\n━━ 4. PER-TENANT ISOLATION")
other = "salesforce.wd12.myworkdayjobs.com"
d3 = browser()
try:
    d3.get(URL)
    d3.execute_script(f"history.replaceState(0,0,'/x'); ")
    h3 = get_handler("workday", None, CFG)
    a1 = h3._account(d3)
    # simulate a different tenant by asking the vault directly
    a2 = CredentialVault(CFG).get_or_create(other, ats="workday")
    check(a1["password"] != a2["password"], "different tenants get different passwords")
finally:
    d3.quit()

print(f"\n━━━━━━ RESULT: {len(fails)} failure(s) ━━━━━━")
for f in fails:
    print("  🐞 " + f)
shutil.rmtree(OUT, ignore_errors=True)
sys.exit(1 if fails else 0)
