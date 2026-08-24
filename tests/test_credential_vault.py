"""
Tests for autonomous ATS self-registration:
  * credential_vault — password generation, per-tenant uniqueness, Excel/CSV
    persistence, reload, status tracking
  * cv_profile — candidate details derived from the CV
  * the six Workday auth bugs these were built to fix

Each Workday test names the bug it pins so a regression is obvious.
"""

import os
import shutil
import sys
import tempfile
import unittest
from unittest.mock import MagicMock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from credential_vault import (
    CredentialVault,
    generate_password,
    password_is_strong,
    plus_address,
    tenant_slug,
)


class TestPasswordGeneration(unittest.TestCase):
    def test_meets_ats_complexity(self):
        for _ in range(40):
            pw = generate_password()
            self.assertGreaterEqual(len(pw), 12)
            self.assertTrue(any(c.islower() for c in pw), pw)
            self.assertTrue(any(c.isupper() for c in pw), pw)
            self.assertTrue(any(c.isdigit() for c in pw), pw)
            self.assertTrue(any(not c.isalnum() for c in pw), pw)
            self.assertTrue(password_is_strong(pw), pw)

    def test_no_symbol_at_the_edges(self):
        # Some ATS validators reject a leading/trailing symbol.
        for _ in range(40):
            pw = generate_password()
            self.assertTrue(pw[0].isalnum(), pw)
            self.assertTrue(pw[-1].isalnum(), pw)

    def test_unique_every_time(self):
        pws = {generate_password() for _ in range(50)}
        self.assertEqual(len(pws), 50)

    def test_length_is_clamped(self):
        self.assertGreaterEqual(len(generate_password(4)), 12)
        self.assertLessEqual(len(generate_password(999)), 64)

    def test_strength_check_rejects_weak(self):
        for weak in ("", "pass", "password", "12345678", "abcdefgh"):
            self.assertFalse(password_is_strong(weak), weak)
        self.assertTrue(password_is_strong("Str0ng!Passw0rd"))


class TestHelpers(unittest.TestCase):
    def test_tenant_slug(self):
        self.assertEqual(tenant_slug("nvidia.wd5.myworkdayjobs.com"), "nvidia")
        self.assertEqual(tenant_slug("careers-acme.icims.com"), "careersacme")

    def test_plus_address(self):
        self.assertEqual(plus_address("me@gmail.com", "nvidia"), "me+nvidia@gmail.com")
        # Never stack tags on an already-tagged address.
        self.assertEqual(plus_address("me+old@gmail.com", "acme"), "me+acme@gmail.com")
        self.assertEqual(plus_address("bad-address", "x"), "bad-address")


class TestVault(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.cfg = {"personal": {"email": "me@example.com"},
                    "credential_vault": {"output_dir": self.dir}}

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def test_mints_and_persists(self):
        v = CredentialVault(self.cfg)
        cred = v.get_or_create("nvidia.wd5.myworkdayjobs.com", ats="workday",
                               job_url="https://x/job/1")
        self.assertEqual(cred["status"], "new")
        self.assertEqual(cred["email"], "me@example.com")
        self.assertTrue(password_is_strong(cred["password"]))
        # A sheet exists on disk (xlsx when openpyxl is present, else csv).
        self.assertTrue(v.xlsx_path.exists() or v.csv_path.exists())

    def test_same_tenant_returns_existing(self):
        v = CredentialVault(self.cfg)
        a = v.get_or_create("nvidia.wd5.myworkdayjobs.com", ats="workday")
        b = v.get_or_create("nvidia.wd5.myworkdayjobs.com", ats="workday")
        self.assertEqual(a["password"], b["password"])   # sign in, don't re-register
        self.assertEqual(b["status"], "existing")

    def test_password_is_unique_per_tenant(self):
        v = CredentialVault(self.cfg)
        a = v.get_or_create("nvidia.wd5.myworkdayjobs.com", ats="workday")
        b = v.get_or_create("salesforce.wd12.myworkdayjobs.com", ats="workday")
        self.assertNotEqual(a["password"], b["password"])

    def test_survives_restart(self):
        v1 = CredentialVault(self.cfg)
        made = v1.get_or_create("acme.wd3.myworkdayjobs.com", ats="workday")
        v2 = CredentialVault(self.cfg)            # fresh process
        again = v2.get_or_create("acme.wd3.myworkdayjobs.com", ats="workday")
        self.assertEqual(made["password"], again["password"])
        self.assertEqual(again["status"], "existing")

    def test_sheet_is_readable_and_complete(self):
        v = CredentialVault(self.cfg)
        v.get_or_create("acme.wd3.myworkdayjobs.com", ats="workday")
        if v.xlsx_path.exists():
            import openpyxl
            ws = openpyxl.load_workbook(v.xlsx_path).active
            headers = [c.value for c in ws[1]]
            row = [c.value for c in ws[2]]
        else:
            import csv
            with open(v.csv_path, newline="", encoding="utf-8") as fh:
                r = list(csv.reader(fh))
            headers, row = r[0], r[1]
        for col in ("site", "email", "password", "status", "created_at"):
            self.assertIn(col, headers)
        self.assertIn("acme.wd3.myworkdayjobs.com", row)

    def test_plus_addressing_optional(self):
        cfg = {**self.cfg, "credential_vault": {**self.cfg["credential_vault"],
                                                "use_plus_addressing": True}}
        v = CredentialVault(cfg)
        cred = v.get_or_create("nvidia.wd5.myworkdayjobs.com", ats="workday")
        self.assertEqual(cred["email"], "me+nvidia@example.com")

    def test_mark_updates_status(self):
        v = CredentialVault(self.cfg)
        v.get_or_create("acme.wd3.myworkdayjobs.com", ats="workday")
        v.mark("acme.wd3.myworkdayjobs.com", "registered")
        self.assertEqual(v.get("acme.wd3.myworkdayjobs.com")["status"], "registered")

    def test_no_email_means_no_credential(self):
        v = CredentialVault({"credential_vault": {"output_dir": self.dir}})
        self.assertEqual(v.get_or_create("x.wd1.myworkdayjobs.com"), {})


class TestWorkdayBugFixes(unittest.TestCase):
    """Each test pins one of the six bugs found by inspecting the auth flow."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.cfg = {"personal": {"email": "me@example.com"},
                    "credential_vault": {"output_dir": self.dir},
                    "external_apply": {"ats_accounts": {}}}
        from ats_handlers import get_handler
        self.h = get_handler("workday", None, self.cfg)

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def _driver(self, host="nvidia.wd5.myworkdayjobs.com"):
        d = MagicMock()
        d.current_url = f"https://{host}/en-US/careers/job/x"
        return d

    def test_bug1_registers_with_no_configured_password(self):
        # BUG 1: previously returned no password and refused to register.
        acct = self.h._account(self._driver())
        self.assertTrue(acct["password"], "no password minted")
        self.assertTrue(password_is_strong(acct["password"]))
        self.assertEqual(acct["status"], "new")

    def test_bug2_credential_is_persisted(self):
        # BUG 2: a generated password used to vanish — the account was unusable.
        self.h._account(self._driver())
        v = CredentialVault(self.cfg)
        self.assertIsNotNone(v.get("nvidia.wd5.myworkdayjobs.com"))

    def test_bug3_password_differs_per_tenant(self):
        a = self.h._account(self._driver("nvidia.wd5.myworkdayjobs.com"))
        b = self.h._account(self._driver("salesforce.wd12.myworkdayjobs.com"))
        self.assertNotEqual(a["password"], b["password"])

    def test_bug4_needs_auth_ignores_page_text(self):
        # BUG 4: "Sign In" in the header after a SUCCESSFUL registration made
        # _needs_auth return True, so success was read as failure.
        d = MagicMock()
        d.find_elements.return_value = []      # no visible auth inputs
        body = MagicMock()
        body.text = "Your account has been created. Sign In | Help | Create Account"
        d.find_element.return_value = body
        self.assertFalse(self.h._needs_auth(d))

    def test_bug4_still_detects_a_real_auth_form(self):
        d = MagicMock()
        field = MagicMock()
        field.is_displayed.return_value = True
        d.find_elements.return_value = [field]
        self.assertTrue(self.h._needs_auth(d))

    def test_bug4_ignores_hidden_auth_widgets(self):
        d = MagicMock()
        hidden = MagicMock()
        hidden.is_displayed.return_value = False
        d.find_elements.return_value = [hidden]
        self.assertFalse(self.h._needs_auth(d))

    def test_bug5_password_hint_is_not_an_error(self):
        # BUG 5: a live password-policy hint aborted account creation.
        d = MagicMock()
        hint = MagicMock()
        hint.is_displayed.return_value = True
        hint.text = "Password must contain at least 1 number and 1 special character"
        hint.get_attribute.return_value = ""
        d.find_elements.return_value = [hint]
        self.assertFalse(self.h._auth_error(d))

    def test_bug5_real_error_still_detected(self):
        d = MagicMock()
        err = MagicMock()
        err.is_displayed.return_value = True
        err.text = "An account with this email already exists"
        err.get_attribute.return_value = "errorMessage"
        d.find_elements.return_value = [err]
        self.assertTrue(self.h._auth_error(d))
        self.assertTrue(self.h._account_exists_error(d))

    def test_bug6_weak_configured_password_is_replaced(self):
        cfg = {**self.cfg, "external_apply": {
            "ats_accounts": {"workday": {"email": "me@example.com", "password": "abc"}}}}
        from ats_handlers import get_handler
        h = get_handler("workday", None, cfg)
        acct = h._account(self._driver())
        self.assertNotEqual(acct["password"], "abc")
        self.assertTrue(password_is_strong(acct["password"]))

    def test_strong_configured_password_is_respected(self):
        cfg = {**self.cfg, "external_apply": {
            "ats_accounts": {"workday": {"email": "me@example.com",
                                         "password": "MyStr0ng!Pass"}}}}
        from ats_handlers import get_handler
        h = get_handler("workday", None, cfg)
        self.assertEqual(h._account(self._driver())["password"], "MyStr0ng!Pass")

    def test_vault_can_be_disabled(self):
        cfg = {**self.cfg, "credential_vault": {"output_dir": self.dir,
                                                "auto_register": False}}
        from ats_handlers import get_handler
        h = get_handler("workday", None, cfg)
        self.assertFalse(h._account(self._driver()).get("password"))


class TestCVProfile(unittest.TestCase):
    CV = """ANKIT KUMAR
Manager, Financial Services Risk Management | ankit.kumar@example.com | +91 98765 43210
Gurugram, India | https://linkedin.com/in/ankitkumar | https://github.com/ankitk

EXPERIENCE (7+ years):
Ernst & Young LLP (Aug 2021-Present): Manager FSRM. Basel III, IRB, RWA, PD/LGD/EAD.
KPMG (2021): Consultant. OFSAA validation, SQL testing.

EDUCATION: PGDM Data Science, IFMR Krea University, 2018
SKILLS: Python, SQL, R, Tableau, Basel III, IRB, Credit Risk, OFSAA
"""

    def test_extracts_the_standard_fields(self):
        from cv_profile import extract_profile
        p = extract_profile(self.CV)
        self.assertEqual(p["first_name"], "Ankit")
        self.assertEqual(p["last_name"], "Kumar")
        self.assertEqual(p["email"], "ankit.kumar@example.com")
        self.assertIn("98765", p["phone"])
        self.assertEqual(p["city"], "Gurugram")
        self.assertEqual(p["country"], "India")
        self.assertIn("linkedin.com/in/ankitkumar", p["linkedin"])
        self.assertIn("github.com/ankitk", p["github"])
        self.assertEqual(p["years_of_experience"], "7")
        self.assertIn("Python", p["skills"])
        self.assertTrue(p["education"])

    def test_enrich_only_fills_blanks(self):
        from cv_profile import enrich_config_profile
        cfg = {"ai": {"cv_text": self.CV},
               "personal": {"first_name": "Preset", "email": ""}}
        out = enrich_config_profile(cfg)
        self.assertEqual(out["personal"]["first_name"], "Preset")     # user wins
        self.assertEqual(out["personal"]["email"], "ankit.kumar@example.com")
        self.assertEqual(out["application"]["years_of_experience"], "7")

    def test_no_cv_is_a_noop(self):
        from cv_profile import enrich_config_profile
        cfg = {"personal": {"first_name": "X"}}
        self.assertEqual(enrich_config_profile(cfg)["personal"]["first_name"], "X")

    def test_garbage_does_not_invent_a_name(self):
        from cv_profile import extract_profile
        p = extract_profile("The quick brown fox jumps over the lazy dog resume")
        self.assertEqual(p["first_name"], "")


if __name__ == "__main__":
    unittest.main()
